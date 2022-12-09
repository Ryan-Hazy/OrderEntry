
# This just imports the libraries that are used during this program
from datetime import datetime
import pandas as pd
import os
import random
from openpyxl import Workbook, load_workbook

# ini moves from one OE sheet to the next
ini=0

#this variable is to stop the while loop that is trying to find the correct sheet name
curs = 0


# This goes through the specified folder where the saved attachments reside
entries = os.listdir(r"C:\Users\hazyr\OneDrive - dematic.com\Attachments")

#This establishes the output workbook
owb = Workbook()
#This is the output workbook sheet that the data is written on
ows = owb.active

#This is printing the collumn headings to the sheet
ows.append(['Change','b', 'Project Number', 'Customer', 'b', 'b', 'Profit center', 'b', 'Salesman', 'b', 'New Business', 'Margin', 'b', 'Category', 'b', 'Finance Charge', 'Cost', 'NBC', 'Check'])
#This is printing the formulas in the next row
ows.append([' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ','=K2-Q2', ' ',' ',' ',' ',' ',' ','=IF(K2=R2, " ", "ERROR")' ])

#each time this loop runs a new OE sheet is being processed
for i in range(len(entries)):

   # curfil is the current file name and that comes from the entries variable above
   curfil = entries[ini]

   # this is to move on to the next sheet if the sheet is a guess number
   si = 0

   # this is the path of the folder where the files reside
   filpat = r"C:\Users\hazyr\OneDrive - dematic.com\Attachments"

   # this combines the file path and name to create a complete address so the file can be used by the program
   filnam=os.path.join(filpat,curfil)

   # This reads the OE File in a way that the worksheet names can be found
   sfinder = pd.ExcelFile(filnam)

   # this reads the file and grabs the OE Sheet file and only the values not any formulas
   oewb = load_workbook(filnam, data_only=True)
   # This reads the actual worksheet
   oews = oewb['OE Sheet']

   # ix is the variable that goes down the lines of the OE Sheet
   ix = str(36)

    # This loop goes down the line of the OE Sheet tab
   for i in range(41):
       t = -1
       # This finds project number
       pronum = oews["B"+ix].value
       # This finds customer name
       cust = oews['F5'].value
       # This finds profit center
       procen = oews["N"+ix].value
       # This finds the sales person
       salper = oews['B12'].value
       # This finds the new business
       newbus = oews["C"+ix].value
       # This finds the category
       category = oews["I"+ix].value
       # This finds whether or not it is a change order
       change = oews['F9'].value

  # this is writing to the output file
       if (newbus is not None):
           try:
               # this loop is trying to match the sheet name to the project number to get the correct sheet
               while (curs == 0):
                   t = t + 1

                   # this is the sheet names
                   snames = str(sfinder.sheet_names[t])

                   # This is asking if the sheet name is the same as the project number
                   if (snames == pronum):

                       # this variable will kill the script when it is equal to one
                       curs = 1
               curs = 0
               # This is the project worksheet
               otherws = oewb.worksheets[t]

               # This finds the cost. This is used to calculate margin
               cost = otherws['B2'].value
               # This finds the finance charge
               finch = otherws['R94'].value
               # This is a check and gets the new business from the project worksheet to see if that is the correct project worksheet
               nbc = otherws['B3'].value
           except:
               otherws = oewb.worksheets[19 + si]
               cost = otherws['B2'].value
               finch = otherws['R94'].value
               nbc = otherws['B3'].value
               si = si + 1

            # This prints out what was found by the program
           ows.append([change, ' ', pronum, cust, ' ', ' ', procen, ' ', salper, ' ', newbus, ' ', ' ', category, ' ', finch, cost, nbc, ' '])

       ix = int(ix) + 1
       ix = str(ix)
   ini = ini + 1


# this is getting the current date to be used for the file name
now = datetime.now()
date = str(now.date())

# this is a random number generator to make sure that files do not get named the same
rn = "-" + str(random.randint(100,999))
fname = date + rn + ".xlsx"

#this section is saving the excel workbook as the file name decided in the previous lines
owb.save(fname)

# this is so you can see what the random number attached to your file is
print(rn)

